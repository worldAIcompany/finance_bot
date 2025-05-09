import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Создаем файл для переводов инвесторов
wb = Workbook()
ws = wb.active
ws.title = "Переводы инвесторов"

# Заголовки
headers = ["Инвестор", "Сумма", "Валюта", "Дата перевода"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Список инвесторов
investors = [
    "Андрей Кобзев",
    "Андрей Несытов",
    "Василий Иванович",
    "Кирилл Бабий",
    "Максим Коптелов",
    "Юрий Костенко",
    "Григорий Перевощиков",
    "Arvind Dhar"
]

# Список валют
currencies = ["USD", "EUR", "RUB", "UAH", "INR", "TRY"]

# Добавляем валидацию данных
# Для инвесторов
dv_investors = DataValidation(type="list", formula1=f'"{",".join(investors)}"')
ws.add_data_validation(dv_investors)
dv_investors.add('A2:A1000')

# Для валют
dv_currencies = DataValidation(type="list", formula1=f'"{",".join(currencies)}"')
ws.add_data_validation(dv_currencies)
dv_currencies.add('C2:C1000')

# Устанавливаем ширину столбцов
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

# Добавляем 20 пустых строк для заполнения
for row in range(2, 22):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

wb.save('investor_transfers.xlsx')

# Создаем файл для оплаты сервисов
wb = Workbook()
ws = wb.active
ws.title = "Оплата сервисов"

# Заголовки
headers = ["Название сервиса", "Сумма", "Валюта", "Дата оплаты", 
          "Период оплаты", "Единица периода", "Кто оплатил"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Список периодов
periods = ["WEEK", "MONTH", "YEAR"]

# Добавляем "ОБЩИЙ" к списку плательщиков
payers = investors + ["ОБЩИЙ"]

# Добавляем валидацию данных
# Для валют
dv_currencies = DataValidation(type="list", formula1=f'"{",".join(currencies)}"')
ws.add_data_validation(dv_currencies)
dv_currencies.add('C2:C1000')

# Для единиц периода
dv_periods = DataValidation(type="list", formula1=f'"{",".join(periods)}"')
ws.add_data_validation(dv_periods)
dv_periods.add('F2:F1000')

# Для плательщиков
dv_payers = DataValidation(type="list", formula1=f'"{",".join(payers)}"')
ws.add_data_validation(dv_payers)
dv_payers.add('G2:G1000')

# Устанавливаем ширину столбцов
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 20

# Добавляем 20 пустых строк для заполнения
for row in range(2, 22):
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

wb.save('service_payments.xlsx') 